import pandas as pd
import argparse
import sys
import os
import zipfile
import openpyxl
from openpyxl.styles import Border, Side, Alignment

# Function to filter specific columns from a CSV file and save the result as an XLSX file
def filter_columns(csv_file, output_file, columns_to_filter):
    try:
        # Load the CSV file into a DataFrame
        df = pd.read_csv(csv_file)
        
        # Filter specific columns
        filtered_df = df[columns_to_filter]
        
        # Create a temporary directory if it doesn't exist
        temp_dir = "temp"
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save the filtered DataFrame to a new XLSX file
        filtered_xlsx_path = os.path.join(temp_dir, output_file)
        filtered_df.to_excel(filtered_xlsx_path, index=False)
        
        return filtered_xlsx_path
    except Exception as e:
        print("An error occurred:", str(e))
        return None

# Function to apply a thin border to cells within specified rows and columns of a worksheet
def apply_border(worksheet, rows, cols):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=rows[0], max_row=rows[1], min_col=cols[0], max_col=cols[1]):
        for cell in row:
            cell.border = thin_border

# Function to apply center and middle alignment to cells within specified rows and columns of a worksheet
def apply_alignment(worksheet, rows, cols):
    alignment = Alignment(horizontal='center', vertical='center')
    
    for row in worksheet.iter_rows(min_row=rows[0], max_row=rows[1], min_col=cols[0], max_col=cols[1]):
        for cell in row:
            cell.alignment = alignment

# Function to create a zip file containing specified files
def create_zip(zip_filename, files_to_zip):
    try:
        with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zipf:
            for file in files_to_zip:
                zipf.write(file, os.path.basename(file))
        print(f"Zip file created: {zip_filename}")
    except Exception as e:
        print("An error occurred:", str(e))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filter specific columns from a CSV file, apply styles, and create a zip archive.")
    parser.add_argument("--input", required=True, help="Input CSV file")
    parser.add_argument("--output", required=True, help="Output XLSX file")
    parser.add_argument("--columns", required=True, help="Columns to filter, comma-separated")
    parser.add_argument("--zip", required=True, help="Zip filename")
    
    args = parser.parse_args()
    
    input_csv_file = args.input
    output_xlsx_file = args.output
    columns_to_filter = args.columns.split(",")
    zip_filename = args.zip
    
    filtered_xlsx_path = filter_columns(input_csv_file, output_xlsx_file, columns_to_filter)
    
    if filtered_xlsx_path:
        try:
            # Load the filtered XLSX file and apply border and alignment styles
            wb = openpyxl.load_workbook(filtered_xlsx_path)
            sheet = wb.active
            
            # Apply a thin border to cells
            apply_border(sheet, (1, sheet.max_row), (1, sheet.max_column))
            
            # Apply center and middle alignment to header cells
            apply_alignment(sheet, (1, 1), (1, sheet.max_column))
            
            # Apply center and middle alignment to all cells
            apply_alignment(sheet, (1, sheet.max_row), (1, sheet.max_column))
            
            # Insert rows and columns to leave the first row and first column empty
            sheet.insert_rows(1)
            sheet.insert_cols(1)
            
            # Set wrap text for all cells
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
            
            # Auto-adjust column widths based on content
            for column in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Save the modified XLSX file
            wb.save(filtered_xlsx_path)
            print("Filtered XLSX file saved.")
        except Exception as e:
            print("An error occurred:", str(e))
        
        try:
            create_zip(zip_filename, [filtered_xlsx_path])
        except Exception as e:
            print("An error occurred:", str(e))
        
        try:
            os.replace(zip_filename, os.path.join(os.path.dirname(os.path.abspath(__file__)), zip_filename))
            print(f"Zip file moved to current directory: {zip_filename}")
        except Exception as e:
            print("An error occurred:", str(e))
