import pandas as pd
import argparse
import sys
import os
import zipfile
import openpyxl
from openpyxl.styles import Border, Side, Alignment

# Function to filter specific columns from a CSV file and save the result as an XLSX file
def filter_columns(csv_file, output_file, columns_to_filter):
    try:
        # Load the CSV file into a DataFrame
        df = pd.read_csv(csv_file)
        
        # Filter specific columns
        filtered_df = df[columns_to_filter]
        
        # Create a temporary directory if it doesn't exist
        temp_dir = "temp"
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save the filtered DataFrame to a new XLSX file
        filtered_xlsx_path = os.path.join(temp_dir, output_file)
        filtered_df.to_excel(filtered_xlsx_path, index=False)
        
        return filtered_xlsx_path
    except Exception as e:
        print("An error occurred:", str(e))
        return None

# Function to apply a thin border to cells within specified rows and columns of a worksheet
def apply_border(worksheet, rows, cols):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows(min_row=rows[0], max_row=rows[1], min_col=cols[0], max_col=cols[1]):
        for cell in row:
            cell.border = thin_border

# Function to create a zip file containing specified files
def create_zip(zip_filename, files_to_zip):
    try:
        with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zipf:
            for file in files_to_zip:
                zipf.write(file, os.path.basename(file))
        print(f"Zip file created: {zip_filename}")
    except Exception as e:
        print("An error occurred:", str(e))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filter specific columns from a CSV file, apply styles, and create a zip archive.")
    parser.add_argument("--input", required=True, help="Input CSV file")
    parser.add_argument("--columns", required=True, help="Columns to filter, comma-separated")
    
    args = parser.parse_args()
    
    input_csv_file = args.input
    base_filename = os.path.splitext(os.path.basename(input_csv_file))[0]
    output_xlsx_file = base_filename + ".xlsx"
    zip_filename = base_filename + ".zip"
    
    filtered_xlsx_path = filter_columns(input_csv_file, output_xlsx_file, args.columns.split(","))
    
    if filtered_xlsx_path:
        try:
            wb = openpyxl.load_workbook(filtered_xlsx_path)
            sheet = wb.active
            
            # Insert an empty row and an empty column before the table
            sheet.insert_rows(1)
            sheet.insert_cols(1)
            
            # Apply border to all cells in the sheet
            apply_border(sheet, (2, sheet.max_row), (2, sheet.max_column))
            
            # Apply alignment to header and content cells
            for row in sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Auto-adjust column widths based on content
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filtered_xlsx_path)
            print("Alignment, border, and column widths adjusted successfully.")
            
            try:
                create_zip(zip_filename, [filtered_xlsx_path])
                os.replace(zip_filename, os.path.join(os.path.dirname(os.path.abspath(__file__)), zip_filename))
                print(f"Zip file created and moved to current directory: {zip_filename}")
            except Exception as e:
                print("An error occurred:", str(e))
            
        except Exception as e:
            print("An error occurred:", str(e))
