import os
import sys
import re
import requests
import zipfile
from io import BytesIO
from openpyxl import load_workbook

domains_extension = "com|org|net"  # Add more extensions if needed

def get_content(input_source):
    
    if input_source.startswith("http://") or input_source.startswith("https://"):
        response = requests.get(input_source)
        if response.status_code == 200:
            return response.text
        else:
            print("Error:", response.status_code)
            sys.exit(1)
    elif os.path.isfile(input_source):
        with open(input_source, "r") as file:
            return file.read()
    else:
        print("Input source not found")
        sys.exit(1)
        

def extract_ips_domains_urls_files(content):
    ip_regex = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    domain_regex = r'\b(?:[a-zA-Z0-9-]+\.)+(?:' + domains_extension + r')\b'
    url_regex = r'http[s]?://(?:[a-zA-Z0-9-]+\.)+(?:' + domains_extension + r')(?:/[a-zA-Z0-9-./?#]+)?'
    file_regex = r'\b(?:[a-zA-Z0-9-]+\.)+(?:txt|pdf|docx?)\b'
    
    ips = re.findall(ip_regex, content)
    domains = re.findall(domain_regex, content)
    urls = re.findall(url_regex, content)
    files = re.findall(file_regex, content)
    
    return ips, domains, urls, files

def convert_excel_to_zip(excel_path):
    wb = load_workbook(excel_path)
    output_zip_path = os.path.splitext(excel_path)[0] + ".zip"
    
    with zipfile.ZipFile(output_zip_path, "w") as zipf:
        # for sheet_name in wb.sheetnames:
        #     sheet = wb[sheet_name]
        #     sheet_data = BytesIO()
        #     sheet.save(sheet_data)
        #     zipf.writestr(sheet_name + ".xlsx", sheet_data.getvalue())
        zipf.write(".\Book2.xlsx")
    
    return output_zip_path

def main():
    if len(sys.argv) != 2:
        print("Usage: python script.py input_source")
        sys.exit(1)
    
    input_source = sys.argv[1]
    content = get_content(input_source)
    ips, domains, urls, files = extract_ips_domains_urls_files(content)
    
    print("IP Addresses:", ips)
    print("Domains:", domains)
    print("URLs:", urls)
    print("Files:", files)
    
    excel_file = ".\Book2.xlsx"
    
    
    #.xlsx"  # Provide the path to your Excel file
    zip_file = convert_excel_to_zip(excel_file)
    print(f"Excel file '{excel_file}' converted and saved as '{zip_file}'")

if __name__ == "__main__":
    main()
